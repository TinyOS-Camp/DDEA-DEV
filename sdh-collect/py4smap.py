import simplejson as json
import StringIO
import pycurl

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

from metadata import *

def setColumnName():
    global ws, wb, dest_file 
    col = 1

    ws.cell('%s%s'%(get_column_letter(col), 1)).value = 'uuid'
    ws.cell('%s%s'%(get_column_letter(col+1), 1)).value = 'Unit'

    if( opt == 'KETI Motes' ):
        ws.cell('%s%s'%(get_column_letter(col+2), 1)).value = 'nodeid'
        ws.cell('%s%s'%(get_column_letter(col+3), 1)).value = 'sensor type'
        ws.cell('%s%s'%(get_column_letter(col+4), 1)).value = 'SerialNumber'
        ws.cell('%s%s'%(get_column_letter(col+5), 1)).value = 'installed info(Height)'
        ws.cell('%s%s'%(get_column_letter(col+6), 1)).value = 'extra info(VAV)'
        #ws.cell('%s%s'%(get_column_letter(col+7), 1)).value = 'location(Building)'
        #ws.cell('%s%s'%(get_column_letter(col+8), 1)).value = 'location(FL)'
        #ws.cell('%s%s'%(get_column_letter(col+9), 1)).value = 'location(Room)'
        col = col + 7

    elif( opt == 'SDH Dent Meters' ):
        ws.cell('%s%s'%(get_column_letter(col+2), 1)).value = 'MeterName'
        ws.cell('%s%s'%(get_column_letter(col+3), 1)).value = 'Model'
        ws.cell('%s%s'%(get_column_letter(col+4), 1)).value = 'installed info(System)'
        ws.cell('%s%s'%(get_column_letter(col+5), 1)).value = 'extra info(System Detail)'
        #ws.cell('%s%s'%(get_column_letter(col+6), 1)).value = 'location(Building)'
        #ws.cell('%s%s'%(get_column_letter(col+7), 1)).value = 'location(FL)'
        #ws.cell('%s%s'%(get_column_letter(col+8), 1)).value = 'location(Room)'
        col = col + 6

    elif( opt == 'Sutardja Dai Hall BACnet' ):
        ws.cell('%s%s'%(get_column_letter(col+2), 1)).value = 'Model'
        ws.cell('%s%s'%(get_column_letter(col+3), 1)).value = 'installed info(Operator)'
        ws.cell('%s%s'%(get_column_letter(col+4), 1)).value = 'extra info(Type)'
        col = col + 5

    ws.cell('%s%s'%(get_column_letter(col), 1)).value = 'location(Building)'
    ws.cell('%s%s'%(get_column_letter(col+1), 1)).value = 'location(FL)'
    ws.cell('%s%s'%(get_column_letter(col+2), 1)).value = 'location(Room)'

def setPyxl(data):
    global opt, ws, wb, dest_file

    wb = Workbook()
    filename = data[0]['Metadata']['SourceName'] + '_Metadata.xlsx'
    dest_file = r'%s' % filename
    ws = wb.worksheets[0]
    #ws = wb.create_sheet()
    ws.title = data[0]['Metadata']['SourceName']
    #ws.title = 'test'

    setColumnName()

    for i in range(len(data)):
        if( opt == 'KETI Motes' ):
            uuid = getUUID(data[i])
            unit = getUnit(data[i])

            nodeid = getNodeid(data[i])
            sensortype = getSensortype(data[i])
            serialnumber = getSerialNumber(data[i])
            installedInfo = getInstalledInfo(data[i])
            addtionalInfo = getAddtionalInfo(data[i])

            locationBD = getLocationBD(data[i])
            locationFL = getLocationFL(data[i])
            locationRM = getLocationRM(data[i])

            ws.cell('%s%s'%(get_column_letter(1), i+2)).value = '%s' % (uuid)
            ws.cell('%s%s'%(get_column_letter(2), i+2)).value = '%s' % (unit)
            ws.cell('%s%s'%(get_column_letter(3), i+2)).value = '%s' % (nodeid)
            ws.cell('%s%s'%(get_column_letter(4), i+2)).value = '%s' % (sensortype)
            ws.cell('%s%s'%(get_column_letter(5), i+2)).value = '%s' % (serialnumber)
            ws.cell('%s%s'%(get_column_letter(6), i+2)).value = '%s' % (installedInfo)
            ws.cell('%s%s'%(get_column_letter(7), i+2)).value = '%s' % (addtionalInfo)
            ws.cell('%s%s'%(get_column_letter(8), i+2)).value = '%s' % (locationBD)
            ws.cell('%s%s'%(get_column_letter(9), i+2)).value = '%s' % (locationFL)
            ws.cell('%s%s'%(get_column_letter(10), i+2)).value = '%s' % (locationRM)
        elif( opt == 'SDH Dent Meters' ):
            uuid = getUUID(data[i])
            unit = getUnit(data[i])

            nodeid = getNodeid(data[i])
            sensortype = getSensortype(data[i])
            installedInfo = getInstalledInfo(data[i])
            addtionalInfo = getAddtionalInfo(data[i])

            locationBD = getLocationBD(data[i])
            locationFL = getLocationFL(data[i])
            locationRM = getLocationRM(data[i])

            ws.cell('%s%s'%(get_column_letter(1), i+2)).value = '%s' % (uuid)
            ws.cell('%s%s'%(get_column_letter(2), i+2)).value = '%s' % (unit)
            ws.cell('%s%s'%(get_column_letter(3), i+2)).value = '%s' % (nodeid)
            ws.cell('%s%s'%(get_column_letter(4), i+2)).value = '%s' % (sensortype)
            ws.cell('%s%s'%(get_column_letter(5), i+2)).value = '%s' % (installedInfo)
            ws.cell('%s%s'%(get_column_letter(6), i+2)).value = '%s' % (addtionalInfo)
            ws.cell('%s%s'%(get_column_letter(7), i+2)).value = '%s' % (locationBD)
            ws.cell('%s%s'%(get_column_letter(8), i+2)).value = '%s' % (locationFL)
            ws.cell('%s%s'%(get_column_letter(9), i+2)).value = '%s' % (locationRM)

        elif( opt == 'Sutardja Dai Hall BACnet' ):
            uuid = getUUID(data[i])
            unit = getUnit(data[i])

            sensortype = getSensortype(data[i])
            installedInfo = getInstalledInfo(data[i])
            addtionalInfo = getAddtionalInfo(data[i])

            locationBD = getLocationBD(data[i])
            locationFL = getLocationFL(data[i])
            locationRM = getLocationRM(data[i])

            ws.cell('%s%s'%(get_column_letter(1), i+2)).value = '%s' % (uuid)
            ws.cell('%s%s'%(get_column_letter(2), i+2)).value = '%s' % (unit)
            ws.cell('%s%s'%(get_column_letter(3), i+2)).value = '%s' % (sensortype)
            ws.cell('%s%s'%(get_column_letter(4), i+2)).value = '%s' % (installedInfo)
            ws.cell('%s%s'%(get_column_letter(5), i+2)).value = '%s' % (addtionalInfo)
            ws.cell('%s%s'%(get_column_letter(6), i+2)).value = '%s' % (locationBD)
            ws.cell('%s%s'%(get_column_letter(7), i+2)).value = '%s' % (locationFL)
            ws.cell('%s%s'%(get_column_letter(8), i+2)).value = '%s' % (locationRM)

        #print uuid, nodeid, sensortype, locationBD, locationFL, locationRM, installedInfo, addtionalInfo
        print uuid

        #ws.cell('F5').value = uuid
        wb.save(filename = dest_file)
    print "%s Done" % len(data)
        

def getUserData():
    global opt

    c = pycurl.Curl()

    if( opt == 'KETI Motes' ):
        c.setopt(pycurl.URL, 'http://new.openbms.org/backend/api/tags/Metadata__SourceName/KETI%20Motes')
    elif( opt == 'SDH Dent Meters' ):
        c.setopt(pycurl.URL, 'http://new.openbms.org/backend/api/tags/Metadata__SourceName/Soda%20Hall%20Dent%20Meters')
    elif( opt == 'Sutardja Dai Hall BACnet' ):
        c.setopt(pycurl.URL, 'http://new.openbms.org/backend/api/tags/Metadata__SourceName/Sutardja%20Dai%20Hall%20BACnet')
    c.setopt(pycurl.HTTPHEADER, ['Accept: application/json'])
    c.setopt(pycurl.VERBOSE, 0)

    contents = StringIO.StringIO()
    c.setopt(pycurl.WRITEFUNCTION, contents.write) 

    c.perform()

    pyobj = json.loads(contents.getvalue())

    setPyxl(pyobj)

def inputNstart():
    global opt

    inputBuff = input('1) KETI Motes\t2)SDH Dent Meters\t3)Sutardja Dai Hall BACnet\n')

    if( inputBuff == 1 ):
        opt = 'KETI Motes'
        getUserData()
    elif( inputBuff == 2 ):
        opt = 'SDH Dent Meters'
        getUserData()
    elif( inputBuff == 3 ):
        opt = 'Sutardja Dai Hall BACnet'
        getUserData()
    else:
        print "wrong input"
    
inputNstart()
